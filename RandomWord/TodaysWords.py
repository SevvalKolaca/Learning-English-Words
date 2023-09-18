import random
import openpyxl
import pandas as pd
import numpy as np
import os
from pathlib import Path
import touch
import xlsxwriter


def version1(fileName):  # Updates are saved by deleting a file and creating a new file.
    # path = 'English_B1_Words.xlsx'
    df = pd.read_excel(fileName)                    # read Excel file

    words_list = df['WORDS'].tolist()               # dataframe to list
    todays_list = random.choices(words_list, k=5)   # random choice from list

    if words_list:                                  # deleting previously selected words
        print("Today's Words: ", todays_list)
        for word in todays_list:
            words_list.pop(words_list.index(word))
    else:
        print("You learned all the words ;)")

    row = 1
    column = 0
    if os.path.isfile(fileName):
        os.remove(fileName)                         # existing file was deleted
        workbook = xlsxwriter.Workbook(fileName)    # creating a new empty file
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', "WORDS")              # added "Words" title

        for words in words_list:                    # write to excel
            worksheet.write(row, column, words)
            row += 1
        workbook.close()
        print("List updated!!")


def version2(fileName):  # Updates are made via the same file
    workbook = openpyxl.load_workbook(fileName)    # open Excel file
    sheet = workbook.active

    df = pd.read_excel(fileName)
    words_list = df['WORDS'].tolist()
    value_list = df['VALUES'].to_list()

    # todays_list = random.choices(words_list, k=5)
    # print("Today's Words: ", todays_list)

    dictionary = dict(zip(words_list,value_list))  # words are keys and values are values ;)

    unlearned_words_list = [word for word, value in dictionary.items() if value == 0]
    if not unlearned_words_list:                   # If the values of 0 are finished, all words have been learned.
        print("You learned all words ;)")
    else:
        todays_list = random.choices(unlearned_words_list, k=5)
        print(todays_list)

    # for word in todays_list:                       
    #     dictionary[word] = 1

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        word = row[0].value
        if word in todays_list:
            row[1].value = 1                       # The value of today's words has been updated to 1

    workbook.save(fileName)
    workbook.close()


path = "deneme.xlsx"
version2(path)
